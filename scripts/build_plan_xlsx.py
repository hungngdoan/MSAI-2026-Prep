"""Build PyTorch_Study_Plan.xlsx from docs/data/study-plan.json.

study-plan.json is the single source of truth. This workbook is a generated,
human-readable export. Workflow: edit the JSON, then run:

    python scripts/build_plan_xlsx.py

Requires openpyxl (see requirements.txt).
"""
import json
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "docs" / "data" / "study-plan.json"
OUT = ROOT / "PyTorch_Study_Plan.xlsx"

plan = json.loads(SRC.read_text(encoding="utf-8"))
meta = plan["meta"]
chapters = plan["chapters"]
tasks = plan["tasks"]
title_by_id = {c["id"]: c["title"] for c in chapters}

HEAD_FILL = PatternFill("solid", fgColor="1B8F87")
HEAD_FONT = Font(bold=True, color="FFFFFF", size=11)
WRAP_TOP = Alignment(wrap_text=True, vertical="top")
TOP = Alignment(vertical="top")
CENTER = Alignment(horizontal="center", vertical="top")
thin = Side(style="thin", color="DDDDDD")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 26


def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


wb = Workbook()

# ---- Overview ----
ws = wb.active
ws.title = "Overview"
total_min = sum(t["minutes"] for t in tasks)
rows = [
    ("PyTorch Study Plan", ""),
    ("", ""),
    ("Title", meta["title"]),
    ("Version", meta["version"]),
    ("Source", meta["source"]),
    ("Window", f'{meta["startDate"]}  ->  {meta["endDate"]}'),
    ("Goal", meta["goal"]),
    ("", ""),
    ("Total tasks", len(tasks)),
    ("Total chapters", sum(1 for c in chapters if c["id"] != "capstone")),
    ("Total study minutes", total_min),
    ("Total study hours", round(total_min / 60, 1)),
    ("", ""),
    ("Note", "Generated from docs/data/study-plan.json. Edit the JSON (the source "
             "of truth), then rerun scripts/build_plan_xlsx.py."),
]
for r, (k, v) in enumerate(rows, start=1):
    a = ws.cell(row=r, column=1, value=k)
    b = ws.cell(row=r, column=2, value=v)
    a.font = Font(bold=True)
    b.alignment = WRAP_TOP
ws.cell(row=1, column=1).font = Font(bold=True, size=16, color="1B8F87")
set_widths(ws, [22, 80])

# ---- Daily Plan ----
ws = wb.create_sheet("Daily Plan")
headers = ["Block", "Chapter", "Date", "Day", "Type", "Focus Topic",
           "Tasks & Description", "Deliverable", "Source", "Sections",
           "Time (min)", "Status"]
ws.append(headers)
for t in tasks:
    ws.append([
        t["week"], title_by_id.get(t["chapter"], t["chapter"]),
        date.fromisoformat(t["date"]), t["day"], t["type"], t["topic"],
        t["description"], t["deliverable"], t["source"], t["sections"],
        t["minutes"], "",
    ])
style_header(ws, len(headers))
set_widths(ws, [7, 24, 12, 11, 9, 30, 72, 34, 18, 18, 11, 12])
last_block, band = None, False
band_fill = PatternFill("solid", fgColor="F2FBFA")
for r in range(2, ws.max_row + 1):
    blk = ws.cell(row=r, column=1).value
    if blk != last_block:
        band = not band
        last_block = blk
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=r, column=c)
        cell.border = BORDER
        cell.alignment = WRAP_TOP if c in (6, 7, 8) else (CENTER if c in (1, 3, 4, 5, 11) else TOP)
        if band:
            cell.fill = band_fill
    ws.cell(row=r, column=3).number_format = "yyyy-mm-dd"

# ---- Chapters ----
ws = wb.create_sheet("Chapters")
headers = ["Block", "ID", "Chapter", "Days", "Start", "End", "Total min", "Milestones", "URL"]
ws.append(headers)
for c in chapters:
    ct = [t for t in tasks if t["chapter"] == c["id"]]
    ds = sorted(t["date"] for t in ct)
    ws.append([
        c["weeks"], c["id"], c["title"], len(ct),
        date.fromisoformat(ds[0]) if ds else None,
        date.fromisoformat(ds[-1]) if ds else None,
        sum(t["minutes"] for t in ct),
        " | ".join(c.get("milestones", [])), c.get("url", ""),
    ])
style_header(ws, len(headers))
set_widths(ws, [7, 10, 30, 7, 12, 12, 11, 60, 52])
for r in range(2, ws.max_row + 1):
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=r, column=c)
        cell.border = BORDER
        cell.alignment = WRAP_TOP if c in (8, 9) else (CENTER if c in (1, 4, 5, 6, 7) else TOP)
    ws.cell(row=r, column=5).number_format = "yyyy-mm-dd"
    ws.cell(row=r, column=6).number_format = "yyyy-mm-dd"

wb.save(OUT)
print(f"wrote {OUT}")
print(f"sheets: {wb.sheetnames}  tasks: {len(tasks)}  chapters: {len(chapters)}  "
      f"minutes: {total_min} ({round(total_min/60,1)} h)")
